"""Generate checked-in sample bank statement fixtures."""

from pathlib import Path

from openpyxl import Workbook


FIXTURES_ROOT = Path(__file__).resolve().parent


def _ensure_parent(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)


def _write_hdfc_statement(path: Path, rows: list[tuple[str, ...]]) -> None:
    _ensure_parent(path)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Statement"

    sheet.append(["HDFC BANK LTD"])
    sheet.append([])
    sheet.append(["Statement of Account"])
    sheet.append(["Account No:", "XXXX1234"])
    sheet.append(["Branch:", "Sample Branch"])
    sheet.append(["Customer Name:", "Fixture User"])
    sheet.append(["Currency:", "INR"])
    sheet.append(["Statement Period:", "01/11/2025 to 30/11/2025"])
    for _ in range(7):
        sheet.append([])

    sheet.append(
        [
            "Date",
            "Narration",
            "Chq./Ref.No.",
            "Value Dt",
            "Withdrawal Amt.",
            "Deposit Amt.",
            "Closing Balance",
        ]
    )
    for row in rows:
        sheet.append(list(row))
    sheet.append([])
    sheet.append(["*** End of Statement ***"])
    workbook.save(path)


def _write_sbi_excel(path: Path, rows: list[tuple[str, ...]]) -> None:
    _ensure_parent(path)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Statement"
    sheet.append(["STATE BANK OF INDIA"])
    sheet.append(["Account Statement"])
    sheet.append(["Account Number:", "XXXX4321"])
    sheet.append(["Branch:", "Sample Branch"])
    sheet.append(["Statement Period:", "01 Nov 2025 to 30 Nov 2025"])
    sheet.append([])
    sheet.append([])
    sheet.append([])
    sheet.append(
        [
            "Txn Date",
            "Value Date",
            "Description",
            "Ref No./Cheque No.",
            "Debit",
            "Credit",
            "Balance",
        ]
    )
    for row in rows:
        sheet.append(list(row))
    sheet.append([])
    sheet.append(["--- End of Statement ---"])
    workbook.save(path)


def _write_sbi_tabsep(path: Path, rows: list[tuple[str, ...]]) -> None:
    _ensure_parent(path)
    lines = [
        "STATE BANK OF INDIA",
        "Account Statement",
        "Account Number:\tXXXX4321",
        "Branch:\tSample Branch",
        "",
        "\t".join(
            [
                "Txn Date",
                "Value Date",
                "Description",
                "Ref No./Cheque No.",
                "Debit",
                "Credit",
                "Balance",
            ]
        ),
    ]
    for row in rows:
        lines.append("\t".join(row))
    lines.append("")
    lines.append("--- End of Statement ---")
    path.write_text("\n".join(lines), encoding="utf-8")


def _write_axis_statement(
    path: Path, headers: list[str], rows: list[tuple[str, ...]], preamble_rows: int
) -> None:
    _ensure_parent(path)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Statement"
    sheet.append(["AXIS BANK"])
    sheet.append(["Transaction Details"])
    sheet.append(["Account No:", "XXXX9876"])
    sheet.append(["Branch:", "Sample Branch"])
    sheet.append(["Currency:", "INR"])
    for _ in range(preamble_rows):
        sheet.append([])
    sheet.append(headers)
    for row in rows:
        sheet.append(list(row))
    sheet.append([])
    sheet.append(["** End **"])
    workbook.save(path)


def main() -> None:
    parser_hdfc_rows = [
        ("01/11/25", "UPI-GROCERY MART-PAY", "REF001", "01/11/25", "1,250.00", "", "48,750.00"),
        ("03/11/25", "TRANSFER TO SBI", "REF002", "03/11/25", "10,000.00", "", "38,750.00"),
        ("05/11/25", "CREDIT CARD PAYMENT HDFC CARD", "REF003", "05/11/25", "5,000.00", "", "33,750.00"),
        ("06/11/25", "NEFT-FREELANCE PAYMENT", "REF004", "06/11/25", "", "25,000.00", "58,750.00"),
    ]
    _write_hdfc_statement(
        FIXTURES_ROOT / "parsers/hdfc/basic_credit_debit/input.xlsx",
        parser_hdfc_rows,
    )
    _write_hdfc_statement(FIXTURES_ROOT / "parsers/hdfc/empty_statement/input.xlsx", [])

    parser_sbi_rows = [
        ("03 Nov 2025", "03 Nov 2025", "TRANSFER FROM HDFC", "TXN001", "", "10,000.00", "55,000.00"),
        ("04 Nov 2025", "04 Nov 2025", "Electricity Bill BESCOM", "TXN002", "1,850.00", "", "53,150.00"),
        ("07 Nov 2025", "07 Nov 2025", "FD Interest Credit", "TXN003", "", "520.00", "53,670.00"),
    ]
    _write_sbi_excel(
        FIXTURES_ROOT / "parsers/sbi/excel_basic/input.xlsx",
        parser_sbi_rows,
    )
    _write_sbi_tabsep(
        FIXTURES_ROOT / "parsers/sbi/tabsep_basic/input.xls",
        parser_sbi_rows,
    )
    _write_sbi_excel(FIXTURES_ROOT / "parsers/sbi/empty_statement/input.xlsx", [])

    parser_axis_rows = [
        ("01-11-2025", "SALARY CREDIT NOV25", "CHQ001", "", "65,000.00", "1,32,900.00"),
        ("06-11-2025", "BILL PAY-BROADBAND", "CHQ002", "1,499.00", "", "1,31,401.00"),
        ("08-11-2025", "UPI/Pay/Restaurant", "CHQ003", "1,875.00", "", "1,29,526.00"),
    ]
    _write_axis_statement(
        FIXTURES_ROOT / "parsers/axis/full_headers/input.xlsx",
        ["Tran Date", "Particulars", "Chq No", "Debit", "Credit", "Balance"],
        parser_axis_rows,
        preamble_rows=5,
    )
    _write_axis_statement(
        FIXTURES_ROOT / "parsers/axis/abbrev_headers/input.xlsx",
        ["Tran Date", "PARTICULARS", "CHQNO", "DR", "CR", "BAL"],
        parser_axis_rows,
        preamble_rows=4,
    )
    _write_axis_statement(
        FIXTURES_ROOT / "parsers/axis/empty_statement/input.xlsx",
        ["Tran Date", "Particulars", "Chq No", "Debit", "Credit", "Balance"],
        [],
        preamble_rows=5,
    )

    pipeline_hdfc_rows = [
        ("01/11/25", "UPI-GROCERY MART-PAY", "REF001", "01/11/25", "1,250.00", "", "48,750.00"),
        ("01/11/25", "UPI-GROCERY MART-PAY", "REF001", "01/11/25", "1,250.00", "", "48,750.00"),
        ("03/11/25", "TRANSFER TO SBI", "REF002", "03/11/25", "10,000.00", "", "38,750.00"),
        ("05/11/25", "CREDIT CARD PAYMENT HDFC CARD", "REF003", "05/11/25", "5,000.00", "", "33,750.00"),
    ]
    pipeline_sbi_rows = [
        ("03 Nov 2025", "03 Nov 2025", "TRANSFER FROM HDFC", "TXN001", "", "10,000.00", "55,000.00"),
        ("04 Nov 2025", "04 Nov 2025", "Electricity Bill BESCOM", "TXN002", "1,850.00", "", "53,150.00"),
    ]
    pipeline_axis_rows = [
        ("01-11-2025", "SALARY CREDIT NOV25", "CHQ001", "", "65,000.00", "1,32,900.00"),
        ("06-11-2025", "BILL PAY-BROADBAND", "CHQ002", "1,499.00", "", "1,31,401.00"),
    ]
    _write_hdfc_statement(
        FIXTURES_ROOT / "pipeline/mixed_basic/input/hdfc.xlsx",
        pipeline_hdfc_rows,
    )
    _write_sbi_tabsep(
        FIXTURES_ROOT / "pipeline/mixed_basic/input/sbi.xls",
        pipeline_sbi_rows,
    )
    _write_axis_statement(
        FIXTURES_ROOT / "pipeline/mixed_basic/input/axis.xlsx",
        ["Tran Date", "Particulars", "Chq No", "Debit", "Credit", "Balance"],
        pipeline_axis_rows,
        preamble_rows=5,
    )


if __name__ == "__main__":
    main()
